from flask import Flask, render_template, jsonify, request
import pandas as pd
import json
from datetime import datetime, timedelta
import os
import numpy as np
import serial
import threading
import time
from collections import deque
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Configuration
EXCEL_FILE = 'new_sensor_data.xlsx'
HISTORICAL_DATA_FILE = 'sensor_data.xlsx'
PORT = '/dev/tty.usbserial-AQ03LYY2' 
BAUDRATE = 115200

# Global variables for real-time data
current_sensor_data = {
    'co2': '-', 'tvoc': '-', 'eco2': '-',
    'pm1': '-', 'pm25': '-', 'pm10': '-',
    'temperature': '-', 'humidity': '-',
    'timestamp': None,
    'last_update': None,
    'last_update_co2': None,
    'last_update_pm': None,
    'connection_status': 'Disconnected'
}

# Thread-safe data storage
data_lock = threading.Lock()
serial_connection = None
serial_thread = None

# Sensor-specific headers for Excel
SENSOR_HEADERS = {
    'scd41': ['Timestamp', 'CO2 (ppm)', 'Temperature (°C)', 'Humidity (%)'],
    'ccs811': ['Timestamp', 'eCO2 (ppm)', 'TVOC (ppb)'],
    'sps30': ['Timestamp', 'PM1.0 (µg/m³)', 'PM2.5 (µg/m³)', 'PM10.0 (µg/m³)']
}

def initialize_excel_file():
    """Initialize Excel file with proper sheets if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        for idx, (sensor, headers) in enumerate(SENSOR_HEADERS.items()):
            if idx == 0:
                ws = wb.active
                ws.title = sensor.upper()
            else:
                ws = wb.create_sheet(title=sensor.upper())
            ws.append(headers)
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

def append_to_sensor_sheet(sensor_name, values):
    """Append sensor data to the appropriate Excel sheet - using exact logic from read_serial.py"""
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet_name = sensor_name.upper()

        if sheet_name not in wb.sheetnames:
            print(f"[WARN] Sheet {sheet_name} missing!")
            return

        ws = wb[sheet_name]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if sensor_name == 'scd41':
            row = [timestamp, values.get('CO2'), values.get('Temperature'), values.get('Humidity')]
        elif sensor_name == 'ccs811':
            row = [timestamp, values.get('eCO2'), values.get('TVOC')]
        elif sensor_name == 'sps30':
            row = [timestamp, values.get('PM1.0'), values.get('PM2.5'), values.get('PM10.0')]
        else:
            print(f"[SKIP] Unknown sensor: {sensor_name}")
            return

        ws.append(row)
        wb.save(EXCEL_FILE)
            
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def update_current_data(sensor_name, values):
    """Update the global current sensor data"""
    global current_sensor_data
    now_str = datetime.now().strftime("%H:%M:%S")
    with data_lock:
        if sensor_name == 'scd41':
            current_sensor_data['co2'] = values.get('CO2', '-')
            current_sensor_data['temperature'] = values.get('Temperature', '-')
            current_sensor_data['humidity'] = values.get('Humidity', '-')
            current_sensor_data['last_update_co2'] = now_str
        elif sensor_name == 'ccs811':
            current_sensor_data['eco2'] = values.get('eCO2', '-')
            current_sensor_data['tvoc'] = values.get('TVOC', '-')
            current_sensor_data['last_update_co2'] = now_str
        elif sensor_name == 'sps30':
            current_sensor_data['pm1'] = values.get('PM1.0', '-')
            current_sensor_data['pm25'] = values.get('PM2.5', '-')
            current_sensor_data['pm10'] = values.get('PM10.0', '-')
            current_sensor_data['last_update_pm'] = now_str
        current_sensor_data['timestamp'] = datetime.now().isoformat()
        current_sensor_data['last_update'] = now_str
        current_sensor_data['connection_status'] = 'Connected'

def serial_reader():
    """Background thread to continuously read from serial port - using exact logic from read_serial.py"""
    global serial_connection, current_sensor_data
    
    try:
        # Start Serial - exactly like read_serial.py
        ser = serial.Serial(PORT, BAUDRATE, timeout=2)
        print(f"Reading from {PORT}...")
        
        with data_lock:
            current_sensor_data['connection_status'] = 'Connected'
        
        # Main Loop - exactly like read_serial.py
        while True:
            try:
                line = ser.readline().decode('utf-8').strip()
                if not line:
                    continue

                try:
                    data = json.loads(line)

                    if "sensor" in data and "data" in data:
                        sensor = data["sensor"]
                        values = data["data"]
                        
                        # Update current data for dashboard
                        update_current_data(sensor, values)
                        
                        # Save to Excel
                        append_to_sensor_sheet(sensor, values)
                        
                        print(f"[{datetime.now()}] Logged data for {sensor.upper()}")
                    elif "error" in data:
                        print(f"[ERROR] {data['error']}")

                except json.JSONDecodeError:
                    print(f"[INVALID JSON] {line}")

            except KeyboardInterrupt:
                print("Exiting...")
                break
                
    except Exception as e:
        print(f"Serial connection error: {e}")
        with data_lock:
            current_sensor_data['connection_status'] = 'Disconnected'

def start_serial_thread():
    """Start the serial reading thread"""
    global serial_thread
    if serial_thread is None or not serial_thread.is_alive():
        serial_thread = threading.Thread(target=serial_reader, daemon=True)
        serial_thread.start()
        print("Serial reading thread started")

def get_air_quality_status():
    """Get overall air quality status based on all sensor thresholds"""
    global current_sensor_data
    
    with data_lock:
        data = current_sensor_data.copy()
    
    # Convert string values to numbers
    co2 = data.get('co2', 0)
    if isinstance(co2, str) and co2 != '-':
        co2 = float(co2)
    elif co2 == '-':
        co2 = 0
        
    pm1 = data.get('pm1', 0)
    if isinstance(pm1, str) and pm1 != '-':
        pm1 = float(pm1)
    elif pm1 == '-':
        pm1 = 0
        
    pm25 = data.get('pm25', 0)
    if isinstance(pm25, str) and pm25 != '-':
        pm25 = float(pm25)
    elif pm25 == '-':
        pm25 = 0
        
    pm10 = data.get('pm10', 0)
    if isinstance(pm10, str) and pm10 != '-':
        pm10 = float(pm10)
    elif pm10 == '-':
        pm10 = 0
        
    tvoc = data.get('tvoc', 0)
    if isinstance(tvoc, str) and tvoc != '-':
        tvoc = float(tvoc)
    elif tvoc == '-':
        tvoc = 0
        
    temperature = data.get('temperature', 0)
    if isinstance(temperature, str) and temperature != '-':
        temperature = float(temperature)
    elif temperature == '-':
        temperature = 0
        
    humidity = data.get('humidity', 0)
    if isinstance(humidity, str) and humidity != '-':
        humidity = float(humidity)
    elif humidity == '-':
        humidity = 0
    
    # Check each parameter and determine worst status
    worst_status = "Good"
    affected_groups = []
    
    # CO₂ thresholds
    if co2 > 1000:
        worst_status = "Bad"
        affected_groups.append("Everyone – may cause drowsiness, reduced focus")
    elif co2 > 800:
        if worst_status == "Good":
            worst_status = "Normal"
        affected_groups.append("Sensitive groups (e.g., elderly, children) may feel discomfort")
    
    # PM1.0 thresholds
    if pm1 > 25:
        worst_status = "Bad"
        affected_groups.append("Everyone – particles can reach deep lungs")
    elif pm1 > 10:
        if worst_status == "Good":
            worst_status = "Normal"
        affected_groups.append("Sensitive groups (e.g., asthmatics) may be at risk")
    
    # PM2.5 thresholds
    if pm25 > 35:
        worst_status = "Bad"
        affected_groups.append("Everyone – especially asthmatics and elderly")
    elif pm25 > 25:
        if worst_status == "Good":
            worst_status = "Normal"
        affected_groups.append("Sensitive groups may feel slight respiratory irritation")
    
    # PM10 thresholds
    if pm10 > 100:
        worst_status = "Bad"
        affected_groups.append("Everyone – lung irritation and allergy risks")
    elif pm10 > 45:
        if worst_status == "Good":
            worst_status = "Normal"
        affected_groups.append("Sensitive individuals may develop allergy-like symptoms")
    
    # TVOC thresholds
    if tvoc > 500:
        worst_status = "Bad"
        affected_groups.append("Everyone – may cause headaches or nausea")
    elif tvoc > 250:
        if worst_status == "Good":
            worst_status = "Normal"
        affected_groups.append("Chemically sensitive people may feel irritation")
    
    # Temperature thresholds
    if temperature < 16 or temperature > 28:
        worst_status = "Bad"
        affected_groups.append("Everyone may feel heat/cold stress")
    elif temperature < 19 or temperature > 25:
        if worst_status == "Good":
            worst_status = "Normal"
        affected_groups.append("Discomfort for sensitive groups")
    
    # Humidity thresholds
    if humidity < 25 or humidity > 70:
        worst_status = "Bad"
        affected_groups.append("Everyone – discomfort or health issues likely")
    elif humidity < 30 or humidity > 60:
        if worst_status == "Good":
            worst_status = "Normal"
        affected_groups.append("Dry skin or mold risk for sensitive groups")
    
    # Determine status color
    if worst_status == "Good":
        status_color = "good"
        status_icon = "✅"
        status_description = "Safe for everyone, No action needed"
    elif worst_status == "Normal":
        status_color = "moderate"
        status_icon = "⚠️"
        status_description = "Some groups may be affected"
    else:  # Bad
        status_color = "unhealthy"
        status_icon = "❌"
        status_description = "Health risks present"
    
    return {
        'status': worst_status,
        'color': status_color,
        'icon': status_icon,
        'description': status_description,
        'affected_groups': affected_groups
    }

def get_detailed_recommendations():
    """Get detailed recommendations based on current sensor values"""
    global current_sensor_data
    
    with data_lock:
        data = current_sensor_data.copy()
    
    recommendations = []
    
    # Convert string values to numbers
    co2 = data.get('co2', 0)
    if isinstance(co2, str) and co2 != '-':
        co2 = float(co2)
    elif co2 == '-':
        co2 = 0
        
    pm1 = data.get('pm1', 0)
    if isinstance(pm1, str) and pm1 != '-':
        pm1 = float(pm1)
    elif pm1 == '-':
        pm1 = 0
        
    pm25 = data.get('pm25', 0)
    if isinstance(pm25, str) and pm25 != '-':
        pm25 = float(pm25)
    elif pm25 == '-':
        pm25 = 0
        
    pm10 = data.get('pm10', 0)
    if isinstance(pm10, str) and pm10 != '-':
        pm10 = float(pm10)
    elif pm10 == '-':
        pm10 = 0
        
    tvoc = data.get('tvoc', 0)
    if isinstance(tvoc, str) and tvoc != '-':
        tvoc = float(tvoc)
    elif tvoc == '-':
        tvoc = 0
        
    temperature = data.get('temperature', 0)
    if isinstance(temperature, str) and temperature != '-':
        temperature = float(temperature)
    elif temperature == '-':
        temperature = 0
        
    humidity = data.get('humidity', 0)
    if isinstance(humidity, str) and humidity != '-':
        humidity = float(humidity)
    elif humidity == '-':
        humidity = 0
    
    # CO₂ recommendations
    if co2 > 1000:
        recommendations.append("🌬️ CO₂ > 1000 ppm: Open windows, increase ventilation or use air exchanger")
    elif co2 > 800:
        recommendations.append("💨 CO₂ 801-1000 ppm: Consider opening windows for ventilation")
    
    # PM1.0 recommendations
    if pm1 > 25:
        recommendations.append("😷 PM1.0 > 25 µg/m³: Use HEPA filter, close windows")
    elif pm1 > 10:
        recommendations.append("🔧 PM1.0 11-25 µg/m³: Improve filtration, monitor")
    
    # PM2.5 recommendations
    if pm25 > 35:
        recommendations.append("😷 PM2.5 > 35 µg/m³: Use air purifier, close windows, avoid physical activity")
    elif pm25 > 25:
        recommendations.append("🔧 PM2.5 26-35 µg/m³: Limit outdoor air entry, monitor levels")
    
    # PM10 recommendations
    if pm10 > 100:
        recommendations.append("🌪️ PM10 > 100 µg/m³: Use air purifier, avoid exposure")
    elif pm10 > 45:
        recommendations.append("🔧 PM10 46-100 µg/m³: Reduce dust-generating activities")
    
    # TVOC recommendations
    if tvoc > 500:
        recommendations.append("🧪 TVOC > 500 ppb: Ventilate room, avoid cleaning agents, use air purifier")
    elif tvoc > 250:
        recommendations.append("🧪 TVOC 251-500 ppb: Avoid scented sprays and synthetic chemicals")
    
    # Temperature recommendations
    if temperature < 16 or temperature > 28:
        recommendations.append("🌡️ Temperature < 16°C or > 28°C: Maintain temperature with AC or heater")
    elif temperature < 19 or temperature > 25:
        recommendations.append("🌡️ Temperature outside 19-25°C: Adjust heater or fan as needed")
    
    # Humidity recommendations
    if humidity < 25 or humidity > 70:
        recommendations.append("💧 Humidity < 25% or > 70%: Control humidity, ventilate, use dryer/dehumidifier")
    elif humidity < 30 or humidity > 60:
        recommendations.append("💧 Humidity outside 30-60%: Use humidifier or ventilate")
    
    # If all parameters are good
    if not recommendations:
        recommendations.append("✅ All parameters are within safe ranges")
        recommendations.append("🌱 Maintain good ventilation for optimal air quality")
    
    return recommendations

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/current-data')
def get_current_data():
    global current_sensor_data
    with data_lock:
        data = current_sensor_data.copy()
    # Check if all real-time values are missing
    all_missing = all(data[k] == '-' for k in ['co2','tvoc','eco2','pm1','pm25','pm10','temperature','humidity'])
    # Get air quality status using new threshold system
    air_quality_status = get_air_quality_status() if not all_missing else {'status': '-', 'color': '', 'icon': '-', 'description': '-', 'affected_groups': []}
    recommendations = get_detailed_recommendations() if not all_missing else ['-']
    response_data = {
        'co2': data['co2'],
        'tvoc': data['tvoc'],
        'eco2': data['eco2'],
        'pm1': data['pm1'],
        'pm25': data['pm25'],
        'pm10': data['pm10'],
        'temperature': data['temperature'],
        'humidity': data['humidity'],
        'timestamp': data['timestamp'],
        'last_update': data['last_update'],
        'last_update_co2': data['last_update_co2'],
        'last_update_pm': data['last_update_pm'],
        'connection_status': data['connection_status'],
        'air_quality_status': air_quality_status,
        'recommendations': recommendations
    }
    return jsonify(response_data)

@app.route('/api/historical-data')
def get_historical_data():
    try:
        # Always use HISTORICAL_DATA_FILE for analysis
        file_to_use = HISTORICAL_DATA_FILE
        # Read from Excel file for historical data
        scd41_df = pd.read_excel(file_to_use, sheet_name='SCD41')
        ccs811_df = pd.read_excel(file_to_use, sheet_name='CCS811') 
        sps30_df = pd.read_excel(file_to_use, sheet_name='SPS30')

        scd41_df['Timestamp'] = pd.to_datetime(scd41_df['Timestamp'])
        ccs811_df['Timestamp'] = pd.to_datetime(ccs811_df['Timestamp'])
        sps30_df['Timestamp'] = pd.to_datetime(sps30_df['Timestamp'])

        combined_df = scd41_df.merge(ccs811_df, on='Timestamp', how='outer', suffixes=('', '_ccs'))
        combined_df = combined_df.merge(sps30_df, on='Timestamp', how='outer', suffixes=('', '_sps'))

        combined_df = combined_df.sort_values('Timestamp')
        combined_df = combined_df.fillna(method='ffill')

        if combined_df.empty:
            return jsonify({'error': 'No data available'})

        # Use all available data (no 24-hour cutoff)
        recent_data = combined_df.copy()

        if recent_data.empty:
            return jsonify({'error': 'No data available'})

        recent_data.set_index('Timestamp', inplace=True)
        hourly_data = recent_data.resample('1h').mean()

        result = {
            'timestamps': [ts.strftime('%H:%M') for ts in hourly_data.index],
            'co2': hourly_data['CO2 (ppm)'].fillna(0).tolist(),
            'tvoc': hourly_data['TVOC (ppb)'].fillna(0).tolist(),
            'eco2': hourly_data['eCO2 (ppm)'].fillna(0).tolist(),
            'pm1': hourly_data['PM1.0 (µg/m³)'].fillna(0).tolist(),
            'pm25': hourly_data['PM2.5 (µg/m³)'].fillna(0).tolist(),
            'pm10': hourly_data['PM10.0 (µg/m³)'].fillna(0).tolist(),
            'temperature': hourly_data['Temperature (°C)'].fillna(0).tolist(),
            'humidity': hourly_data['Humidity (%)'].fillna(0).tolist()
        }

        return jsonify(result)
        
    except Exception as e:
        print(f"Error reading historical data: {e}")
        return jsonify({'error': 'No data available'})

@app.route('/api/insights')
def get_insights():
    try:
        # Always use HISTORICAL_DATA_FILE for analysis
        file_to_use = HISTORICAL_DATA_FILE
        # Read from Excel file for insights
        scd41_df = pd.read_excel(file_to_use, sheet_name='SCD41')
        ccs811_df = pd.read_excel(file_to_use, sheet_name='CCS811') 
        sps30_df = pd.read_excel(file_to_use, sheet_name='SPS30')

        scd41_df['Timestamp'] = pd.to_datetime(scd41_df['Timestamp'])
        ccs811_df['Timestamp'] = pd.to_datetime(ccs811_df['Timestamp'])
        sps30_df['Timestamp'] = pd.to_datetime(sps30_df['Timestamp'])

        combined_df = scd41_df.merge(ccs811_df, on='Timestamp', how='outer', suffixes=('', '_ccs'))
        combined_df = combined_df.merge(sps30_df, on='Timestamp', how='outer', suffixes=('', '_sps'))

        combined_df = combined_df.sort_values('Timestamp')
        combined_df = combined_df.fillna(method='ffill')

        if combined_df.empty:
            return jsonify({'error': 'No data available'})

        # Use all available data (no 24-hour cutoff)
        recent_data = combined_df.copy()

        if recent_data.empty:
            return jsonify({'error': 'No data available'})

        insights = {}
        # Helper to check if column exists and has data
        def valid_col(col):
            return col in recent_data.columns and recent_data[col].dropna().shape[0] > 0

        # Helper to get time string from index or column
        def get_time_str(df, idx):
            # idx is the index of the row (could be int or datetime)
            if isinstance(df.index, pd.DatetimeIndex):
                try:
                    return df.index[idx].strftime('%H:%M')
                except Exception:
                    return str(df.index[idx])
            # fallback: try Timestamp column
            ts = df.iloc[idx]['Timestamp'] if 'Timestamp' in df.columns else None
            if pd.notnull(ts):
                try:
                    return pd.to_datetime(ts).strftime('%H:%M')
                except Exception:
                    return str(ts)
            return str(idx)

        # CO2
        if valid_col('CO2 (ppm)'):
            co2_data = recent_data['CO2 (ppm)'].dropna()
            co2_max = co2_data.max()
            co2_min = co2_data.min()
            co2_avg = co2_data.mean()
            max_idx = co2_data.idxmax()
            min_idx = co2_data.idxmin()
            co2_max_time = get_time_str(recent_data, max_idx)
            co2_min_time = get_time_str(recent_data, min_idx)
            co2_threshold = 1000
            co2_threshold_exceeded = (co2_data > co2_threshold).sum()
            insights['co2'] = {
                'max': f"{co2_max:.0f} ppm",
                'max_time': co2_max_time,
                'min': f"{co2_min:.0f} ppm",
                'min_time': co2_min_time,
                'avg': f"{co2_avg:.0f} ppm",
                'threshold': co2_threshold,
                'threshold_exceeded': int(co2_threshold_exceeded)
            }
        # PM1.0
        if valid_col('PM1.0 (µg/m³)'):
            pm1_data = recent_data['PM1.0 (µg/m³)'].dropna()
            pm1_max = pm1_data.max()
            pm1_min = pm1_data.min()
            pm1_avg = pm1_data.mean()
            max_idx = pm1_data.idxmax()
            min_idx = pm1_data.idxmin()
            pm1_max_time = get_time_str(recent_data, max_idx)
            pm1_min_time = get_time_str(recent_data, min_idx)
            pm1_threshold = 25
            pm1_threshold_exceeded = (pm1_data > pm1_threshold).sum()
            insights['pm1'] = {
                'max': f"{pm1_max:.1f} µg/m³",
                'max_time': pm1_max_time,
                'min': f"{pm1_min:.1f} µg/m³",
                'min_time': pm1_min_time,
                'avg': f"{pm1_avg:.1f} µg/m³",
                'threshold': pm1_threshold,
                'threshold_exceeded': int(pm1_threshold_exceeded)
            }
        # PM2.5
        if valid_col('PM2.5 (µg/m³)'):
            pm25_data = recent_data['PM2.5 (µg/m³)'].dropna()
            pm25_max = pm25_data.max()
            pm25_min = pm25_data.min()
            pm25_avg = pm25_data.mean()
            max_idx = pm25_data.idxmax()
            min_idx = pm25_data.idxmin()
            pm25_max_time = get_time_str(recent_data, max_idx)
            pm25_min_time = get_time_str(recent_data, min_idx)
            pm25_threshold = 35
            pm25_threshold_exceeded = (pm25_data > pm25_threshold).sum()
            insights['pm25'] = {
                'max': f"{pm25_max:.1f} µg/m³",
                'max_time': pm25_max_time,
                'min': f"{pm25_min:.1f} µg/m³",
                'min_time': pm25_min_time,
                'avg': f"{pm25_avg:.1f} µg/m³",
                'threshold': pm25_threshold,
                'threshold_exceeded': int(pm25_threshold_exceeded)
            }
        # PM10
        if valid_col('PM10.0 (µg/m³)'):
            pm10_data = recent_data['PM10.0 (µg/m³)'].dropna()
            pm10_max = pm10_data.max()
            pm10_min = pm10_data.min()
            pm10_avg = pm10_data.mean()
            max_idx = pm10_data.idxmax()
            min_idx = pm10_data.idxmin()
            pm10_max_time = get_time_str(recent_data, max_idx)
            pm10_min_time = get_time_str(recent_data, min_idx)
            pm10_threshold = 100
            pm10_threshold_exceeded = (pm10_data > pm10_threshold).sum()
            insights['pm10'] = {
                'max': f"{pm10_max:.1f} µg/m³",
                'max_time': pm10_max_time,
                'min': f"{pm10_min:.1f} µg/m³",
                'min_time': pm10_min_time,
                'avg': f"{pm10_avg:.1f} µg/m³",
                'threshold': pm10_threshold,
                'threshold_exceeded': int(pm10_threshold_exceeded)
            }
        # TVOC
        if valid_col('TVOC (ppb)'):
            tvoc_data = recent_data['TVOC (ppb)'].dropna()
            tvoc_max = tvoc_data.max()
            tvoc_min = tvoc_data.min()
            tvoc_avg = tvoc_data.mean()
            max_idx = tvoc_data.idxmax()
            min_idx = tvoc_data.idxmin()
            tvoc_max_time = get_time_str(recent_data, max_idx)
            tvoc_min_time = get_time_str(recent_data, min_idx)
            tvoc_threshold = 500
            tvoc_threshold_exceeded = (tvoc_data > tvoc_threshold).sum()
            insights['tvoc'] = {
                'max': f"{tvoc_max:.0f} ppb",
                'max_time': tvoc_max_time,
                'min': f"{tvoc_min:.0f} ppb",
                'min_time': tvoc_min_time,
                'avg': f"{tvoc_avg:.0f} ppb",
                'threshold': tvoc_threshold,
                'threshold_exceeded': int(tvoc_threshold_exceeded)
            }
        # Temperature
        if valid_col('Temperature (°C)'):
            temp_data = recent_data['Temperature (°C)'].dropna()
            temp_max = temp_data.max()
            temp_min = temp_data.min()
            temp_avg = temp_data.mean()
            max_idx = temp_data.idxmax()
            min_idx = temp_data.idxmin()
            temp_max_time = get_time_str(recent_data, max_idx)
            temp_min_time = get_time_str(recent_data, min_idx)
            temp_threshold_high = 28
            temp_threshold_low = 16
            temp_threshold_exceeded = ((temp_data > temp_threshold_high) | (temp_data < temp_threshold_low)).sum()
            insights['temperature'] = {
                'max': f"{temp_max:.1f} °C",
                'max_time': temp_max_time,
                'min': f"{temp_min:.1f} °C",
                'min_time': temp_min_time,
                'avg': f"{temp_avg:.1f} °C",
                'threshold_high': temp_threshold_high,
                'threshold_low': temp_threshold_low,
                'threshold_exceeded': int(temp_threshold_exceeded)
            }
        # Humidity
        if valid_col('Humidity (%)'):
            hum_data = recent_data['Humidity (%)'].dropna()
            hum_max = hum_data.max()
            hum_min = hum_data.min()
            hum_avg = hum_data.mean()
            max_idx = hum_data.idxmax()
            min_idx = hum_data.idxmin()
            hum_max_time = get_time_str(recent_data, max_idx)
            hum_min_time = get_time_str(recent_data, min_idx)
            hum_threshold_high = 70
            hum_threshold_low = 25
            hum_threshold_exceeded = ((hum_data > hum_threshold_high) | (hum_data < hum_threshold_low)).sum()
            insights['humidity'] = {
                'max': f"{hum_max:.1f} %",
                'max_time': hum_max_time,
                'min': f"{hum_min:.1f} %",
                'min_time': hum_min_time,
                'avg': f"{hum_avg:.1f} %",
                'threshold_high': hum_threshold_high,
                'threshold_low': hum_threshold_low,
                'threshold_exceeded': int(hum_threshold_exceeded)
            }
        return jsonify(insights)
        
    except Exception as e:
        print(f"Error reading insights: {e}")
        return jsonify({'error': 'No data available'})

if __name__ == '__main__':
    # Initialize Excel file
    initialize_excel_file()
    
    # Start serial reading thread
    start_serial_thread()
    
    print("Flask app starting with real-time serial data integration...")
    app.run(debug=True, port=5000)
